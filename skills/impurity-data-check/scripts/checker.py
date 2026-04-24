#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HPLC有关物质检测数据明显错误检查器
版本: 0.9.4

架构：
1. 提取阶段：脚本从 Excel 中提取项目关键信息（项目编号、药品名、规格等）
2. 判断阶段：Agent 根据提取的信息 + 配置文件，综合判断项目和阈值
3. 检查阶段：基于判断结果执行批号格式 + 忽略不计检查

阈值优先级（从高到低）：
1. --threshold 命令行参数
2. --agent-decision Agent 决策参数
3. 项目级配置 project_configs（项目编号/文件名/内容匹配）
4. 全局默认配置 default_threshold / impurity_thresholds
5. 内置默认值 0.05%
"""

import re
import os
import sys
import yaml
import json
import argparse
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, asdict
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────
# 数据结构
# ─────────────────────────────────────────────────────────────────
@dataclass
class CheckError:
    cell_refs: str
    field: str
    current_values: str
    error_type: str
    message: str
    suggestion: str


@dataclass
class ProjectInfo:
    """从 Excel 中提取的项目信息"""
    project_code: str        # 项目编号，如 LMS002
    drug_name: str          # 药品名称，如 黄体酮
    product_name: str        # 产品名称，如 黄体酮预灌封注射液
    specification: str       # 规格
    batch_range: str         # 批号范围
    detection_type: str      # 检测项目，如 有关物质
    raw_text: str            # 原始关键文本（用于 AI 判断）

    def to_dict(self) -> dict:
        return asdict(self)

    def summary(self) -> str:
        parts = []
        if self.project_code:
            parts.append(f"项目编号: {self.project_code}")
        if self.drug_name:
            parts.append(f"药品名: {self.drug_name}")
        if self.product_name:
            parts.append(f"产品名: {self.product_name}")
        if self.detection_type:
            parts.append(f"检测项目: {self.detection_type}")
        return ", ".join(parts) if parts else "未识别到项目信息"


# ─────────────────────────────────────────────────────────────────
# 项目信息提取器
# ─────────────────────────────────────────────────────────────────
class ProjectInfoExtractor:
    """
    从 Excel 文件中提取项目关键信息
    支持从文件名、表头内容等多个维度提取
    """

    # 常见药品名/通用名模式
    DRUG_NAME_PATTERNS = [
        r'([^\\/\s]+?)(?:预灌封|注射液|片剂|胶囊|颗粒|软膏|凝胶)',
        '([\u4e00-\u9fa5]{2,6})(?:注射液|预灌封)',
        '^[一-龥]+',
    ]

    # 项目编号模式
    PROJECT_CODE_PATTERN = r'\b([A-Z]{2,}\d{3,})\b'
    PROJECT_CODE_PATTERNS = [
        r'\b(LM\w{2,})\b',
        r'\b(PM\w{2,})\b',
        r'\b(XM\w{2,})\b',
        r'\b([A-Z]{2,}\d{3,})\b',
    ]
    def __init__(self):
        self.project_code = ''
        self.drug_name = ''
        self.product_name = ''
        self.specification = ''
        self.batch_range = ''
        self.detection_type = ''

    def extract(self, file_path: str, ws) -> ProjectInfo:
        """从文件路径和 Excel 内容提取项目信息"""
        excel_name = os.path.basename(file_path)
        excel_dir = os.path.dirname(os.path.abspath(file_path))

        # 1. 从文件路径提取
        self._extract_from_path(excel_name, excel_dir)

        # 2. 从 Excel 表头内容提取（前35行）
        self._extract_from_content(ws)

        # 3. 汇总原始文本（用于 AI 判断）
        raw_parts = []
        if excel_name:
            raw_parts.append(f"文件名: {excel_name}")
        if self.project_code:
            raw_parts.append(f"项目编号: {self.project_code}")
        if self.drug_name:
            raw_parts.append(f"药品名: {self.drug_name}")
        if self.product_name:
            raw_parts.append(f"产品名: {self.product_name}")
        if self.detection_type:
            raw_parts.append(f"检测项目: {self.detection_type}")
        if self.batch_range:
            raw_parts.append(f"批号范围: {self.batch_range}")
        raw_text = "\\n".join(raw_parts)

        return ProjectInfo(
            project_code=self.project_code,
            drug_name=self.drug_name,
            product_name=self.product_name,
            specification=self.specification,
            batch_range=self.batch_range,
            detection_type=self.detection_type,
            raw_text=raw_text
        )

    def _extract_from_path(self, filename: str, dir_path: str):
        """从文件路径提取关键信息"""
        # 提取项目编号（如 LMS002）
        for pattern in self.PROJECT_CODE_PATTERNS:
            m = re.search(pattern, filename + "/" + dir_path)
            if m:
                self.project_code = m.group(1)
                break

        # 提取药品名
        drug_match = re.search('([\u4e00-\u9fa5]{2,6}(?:预灌封|注射液|片剂|胶囊))', filename)
        if drug_match:
            self.product_name = drug_match.group(1)
            # 简单提取通用名
            for kw in ['预灌封', '注射液', '片剂', '胶囊', '颗粒']:
                if kw in self.product_name:
                    self.drug_name = self.product_name.split(kw)[0]
                    break

    def _extract_from_content(self, ws):
        """从 Excel 表头内容提取"""
        content_rows = []
        for row in ws.iter_rows(min_row=1, max_row=35):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    content_rows.append(cell.value.strip())

        full_text = "\\n".join(content_rows)

        # 提取检测项目（精确匹配关键字）
        detection_exact = ['有关物质', '含量', '水分', 'PH', '澄明度', '细菌内毒素',
                           '有关物质检测结果', '含量测定', '水分测定']
        matched_kw = None
        for kw in detection_exact:
            if kw in full_text:
                matched_kw = kw
                break
        if matched_kw:
            for suffix in ['检测结果', '测定', '检测']:
                if matched_kw.endswith(suffix):
                    self.detection_type = matched_kw[:-len(suffix)]
                    break
            if not self.detection_type:
                self.detection_type = matched_kw

        # 提取批号范围（A1 标题 或 H3）
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    v = cell.value.strip()
                    if '批号' in v or self.project_code in v:
                        continue
                    m = re.search(rf'{self.project_code}[^\\s,\\`~·]{5,}', v) if self.project_code else None
                    if m:
                        self.batch_range = m.group(0)[:30]
                        break

        # 提取药品名（A1 标题）
        if not self.product_name and content_rows:
            first = content_rows[0]
            m = re.search('([\u4e00-\u9fa5]{2,}(?:预灌封|注射液|片剂|胶囊))', first)
            if m:
                self.product_name = m.group(1)
                for kw in ['预灌封', '注射液']:
                    if kw in self.product_name:
                        self.drug_name = self.product_name.split(kw)[0]
                        break

    def to_prompt_context(self) -> str:
        """生成用于 AI 判断的上下文信息"""
        lines = ["【项目信息提取结果】"]
        if self.project_code:
            lines.append(f"  项目编号: {self.project_code}")
        if self.drug_name:
            lines.append(f"  药品名（通用名）: {self.drug_name}")
        if self.product_name:
            lines.append(f"  产品名（全称）: {self.product_name}")
        if self.specification:
            lines.append(f"  规格: {self.specification}")
        if self.detection_type:
            lines.append(f"  检测项目: {self.detection_type}")
        if self.batch_range:
            lines.append(f"  批号范围: {self.batch_range}")
        lines.append("")
        return "\\n".join(lines)


# ─────────────────────────────────────────────────────────────────
# 配置文件加载器（保持向后兼容）
# ─────────────────────────────────────────────────────────────────
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def find_config_file(excel_path: str) -> Optional[str]:
    skill_config = os.path.join(SKILL_DIR, 'impurity-checker.yaml')
    if os.path.exists(skill_config):
        return skill_config
    return None


def load_config(excel_path: str) -> Dict:
    config_path = find_config_file(excel_path)
    if not config_path:
        return {}
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        return config or {}
    except Exception as e:
        print(f"  ⚠️ 配置文件读取失败: {e}")
        return {}


# ─────────────────────────────────────────────────────────────────
# 阈值解析器（供 Agent 判断用）
# ─────────────────────────────────────────────────────────────────
class ThresholdResolver:
    def __init__(self, config: Dict, project_cfg: Optional[Dict] = None):
        self.config = config
        self.project_cfg = project_cfg
        self.default_threshold = float(
            config.get('default_threshold') or config.get('report_threshold') or 0.05
        )
        # 合并全局和项目级杂质阈值，项目级优先
        merged = dict(config.get('impurity_thresholds') or {})
        if project_cfg:
            merged.update(project_cfg.get('impurity_thresholds') or {})
        self.impurity_thresholds = merged

    def get_threshold(self, impurity_name: str) -> float:
        if not impurity_name or not self.impurity_thresholds:
            return self.default_threshold
        for key, val in self.impurity_thresholds.items():
            if key == impurity_name:
                return float(val)
        for key, val in self.impurity_thresholds.items():
            if key in impurity_name:
                return float(val)
        return self.default_threshold

    def get_default_threshold(self) -> float:
        if self.project_cfg:
            return float(self.project_cfg.get('default_threshold', self.default_threshold))
        return self.default_threshold

    def resolve(self, excel_path: str, active_project: Optional[str] = None) -> Tuple[float, List[str]]:
        lines = []
        config_path = find_config_file(excel_path)
        if config_path:
            lines.append(f"📎 配置文件: {config_path}")
        if active_project and self.project_cfg:
            lines.append(f"📎 匹配项目: {active_project}")
            lines.append(f"📎 项目默认报告限: {self.project_cfg.get('default_threshold', self.default_threshold)}%")
            proj_impurities = self.project_cfg.get('impurity_thresholds', {})
            if proj_impurities:
                lines.append(f"📎 项目杂质阈值: {len(proj_impurities)} 项")
                for name, val in list(proj_impurities.items())[:5]:
                    lines.append(f"    · {name} → {val}%")
        else:
            lines.append(f"📎 默认报告限: {self.default_threshold}%")
            if self.impurity_thresholds:
                lines.append(f"📎 杂质专用阈值: {len(self.impurity_thresholds)} 项")
                for name, val in list(self.impurity_thresholds.items())[:5]:
                    lines.append(f"    · {name} → {val}%")
        return self.get_default_threshold(), lines

    def auto_match_project(self, project_info: ProjectInfo) -> Tuple[Optional[str], Optional[Dict]]:
        """
        根据提取的项目信息，自动匹配配置文件中的项目
        匹配规则：精确匹配 > 包含匹配（第一个匹配项）
        """
        project_configs = self.config.get('project_configs', {})
        if not project_configs:
            return None, None

        # 优先使用项目编号匹配
        if project_info.project_code:
            for keyword, proj_cfg in project_configs.items():
                if keyword == project_info.project_code:
                    return keyword, proj_cfg if isinstance(proj_cfg, dict) else None

        # 使用产品名匹配
        if project_info.product_name:
            for keyword, proj_cfg in project_configs.items():
                if keyword in project_info.product_name or project_info.product_name in keyword:
                    return keyword, proj_cfg if isinstance(proj_cfg, dict) else None

        # 使用药品名匹配
        if project_info.drug_name:
            for keyword, proj_cfg in project_configs.items():
                if keyword in project_info.drug_name or project_info.drug_name in keyword:
                    return keyword, proj_cfg if isinstance(proj_cfg, dict) else None

        # 遍历所有配置，找包含关系
        for keyword, proj_cfg in project_configs.items():
            if (keyword in project_info.raw_text or
                project_info.project_code and keyword in project_info.project_code):
                return keyword, proj_cfg if isinstance(proj_cfg, dict) else None

        return None, None


# ─────────────────────────────────────────────────────────────────
# 批号格式验证器
# ─────────────────────────────────────────────────────────────────
class BatchFormatValidator:
    def __init__(self,
                 pattern: str = r'^(.+)-(\d{2})(\d{2})(\d{2})-(\d{2})$',
                 pattern_desc: str = '项目编号-年月日-序号（例：LMS002-260401-05）'):
        self.pattern = pattern
        self.pattern_desc = pattern_desc
        self._compiled = re.compile(pattern)

    def is_valid(self, batch: str) -> bool:
        if not batch:
            return False
        return self._compiled.match(batch.strip()) is not None

    def normalize(self, batch: str) -> Optional[str]:
        batch = batch.strip()
        if self.is_valid(batch):
            return batch
        match = re.match(r'^(.+)-(\d{8})$', batch)
        if match:
            prefix, digits = match.group(1), match.group(2)
            if digits.isdigit():
                return f"{prefix}-{digits[:6]}-{digits[6:8]}"
        return None


# ─────────────────────────────────────────────────────────────────
# 合并单元格解析器
# ─────────────────────────────────────────────────────────────────
class MergedCellResolver:
    def __init__(self, ws):
        self.ws = ws
        self.master_map: Dict[str, str] = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.bounds
            master_ref = f"{get_column_letter(min_col)}{min_row}"
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    self.master_map[f"{get_column_letter(c)}{r}"] = master_ref

    def get_master(self, cell_ref: str) -> str:
        return self.master_map.get(cell_ref, cell_ref)

    def is_merged(self, cell_ref: str) -> bool:
        return self.get_master(cell_ref) != cell_ref

    def get_merged_range_cells(self, master_ref: str) -> List[str]:
        return [ref for ref, m in self.master_map.items() if m == master_ref]

    def format_cell_ref(self, cell_ref: str) -> str:
        master = self.get_master(cell_ref)
        if self.is_merged(cell_ref):
            merged_cells = self.get_merged_range_cells(master)
            return self._format_range(merged_cells)
        return cell_ref

    def _format_range(self, cell_refs: List[str]) -> str:
        if not cell_refs:
            return ""
        parsed = []
        for ref in cell_refs:
            m = re.match(r'([A-Z]+)(\\d+)', ref)
            if m:
                parsed.append((m.group(1), int(m.group(2))))
        if not parsed:
            return ", ".join(cell_refs)
        row_groups: Dict[int, List[str]] = defaultdict(list)
        for col, row in parsed:
            row_groups[row].append(col)
        parts = []
        for row in sorted(row_groups.keys()):
            cols = sorted(row_groups[row], key=lambda c: ord(c[0]))
            if len(cols) == 1:
                parts.append(f"{cols[0]}{row}")
            elif self._is_consecutive(cols):
                parts.append(f"{cols[0]}-{cols[-1]}/{row}")
            else:
                parts.append(f"{','.join(cols)}/{row}")
        return ", ".join(parts)

    def _is_consecutive(self, cols: List[str]) -> bool:
        nums = sorted([ord(c) - ord('A') for c in cols])
        return all(nums[i+1] - nums[i] == 1 for i in range(len(nums) - 1))


# ─────────────────────────────────────────────────────────────────
# 批号格式检查器
# ─────────────────────────────────────────────────────────────────
class BatchNumberChecker:
    NON_BATCH_KEYWORDS = [
        '对照品', '系统适应性', '灵敏度', '流动相', '稀释', '空白',
        '溶剂', '配制', '来源', '对照', '自身', '杂质', '名称',
        '供试品名称', '批号', '样品', '实验日期', '检验人', '复核人',
        '实验员', '检测项目', '数据存储路径', '方法', '仪器',
        '泵模块', '进样器模块', '柱温箱模块', '检测器模块',
        '理论塔板数', '拖尾因子', '分离度', 'RSD', 'RSD%',
        '峰面积', '保留时间', '相对保留时间', '信噪比',
        '供试品', '灵敏度溶液', '系统适用性', '计算公式',
    ]

    def __init__(self, batch_format: BatchFormatValidator):
        self.batch_format = batch_format
        self.detected: List[Tuple[str, str]] = []

    def _strip_suffix(self, value: str) -> str:
        for suffix in ['（复测）', '(复测)', '（再测）', '(再测)']:
            if suffix in value:
                return value.split(suffix)[0].strip()
        return value.strip()

    def _is_batch_like(self, value: str) -> bool:
        if not value or not isinstance(value, str):
            return False
        value = value.strip()
        for kw in self.NON_BATCH_KEYWORDS:
            if kw in value:
                return False
        if re.match(r'^\\d+$', value):
            return False
        if '://' in value or ('/' in value and not '-' in value):
            return False
        if not any(c.isalpha() for c in value):
            return False
        return True

    def check(self, cell_ref: str, value: str):
        if self._is_batch_like(value):
            self.detected.append((value, cell_ref))

    def final_check(self, resolver: MergedCellResolver) -> List[CheckError]:
        errors = []
        seen_masters = set()
        for raw_value, cell_ref in self.detected:
            batch_body = self._strip_suffix(raw_value)
            if self.batch_format.is_valid(batch_body):
                continue
            normalized = self.batch_format.normalize(batch_body)
            if not normalized:
                continue
            if batch_body != raw_value:
                normalized += raw_value[len(batch_body):]
            master = resolver.get_master(cell_ref)
            is_merged = resolver.is_merged(cell_ref)
            if is_merged:
                if master in seen_masters:
                    continue
                seen_masters.add(master)
                formatted = resolver.format_cell_ref(cell_ref)
            else:
                formatted = resolver.format_cell_ref(cell_ref)

            errors.append(CheckError(
                cell_refs=formatted,
                field='批号格式',
                current_values=raw_value,
                error_type='BATCH_FORMAT_INVALID',
                message=f'批号格式不符合标准（{self.batch_format.pattern_desc}）',
                suggestion=normalized
            ))
        return errors


# ─────────────────────────────────────────────────────────────────
# 忽略不计检查器
# ─────────────────────────────────────────────────────────────────
class IgnoreTermChecker:
    def __init__(self, threshold_resolver: ThresholdResolver):
        self.resolver = threshold_resolver
        self.IGNORE_KW = ['忽略不计', '忽略', '不及']
        self.detected: List[Tuple[str, str, str]] = []

    def check(self, content_value, report_value: str, cell_ref: str, impurity_name: str = ''):
        is_ignored = any(kw in str(report_value) for kw in self.IGNORE_KW)
        if is_ignored:
            self.detected.append((str(content_value), cell_ref, impurity_name))

    def final_check(self, resolver):
        """检查忽略不计标记错误的最终报告"""
        errors = []
        seen = set()
        for content_str, cell_ref, impurity_name in self.detected:
            # 解析数值
            m = re.search(r'[\d.]+', str(content_str))
            if not m:
                continue
            try:
                num_val = float(m.group())
            except ValueError:
                continue
            # 获取阈值
            threshold = self.resolver.get_threshold(impurity_name)
            # 判断是否应标记错误
            if num_val < threshold:
                continue
            # 处理合并单元格
            master = resolver.get_master(cell_ref)
            is_merged = resolver.is_merged(cell_ref)
            if is_merged:
                if master in seen:
                    continue
                seen.add(master)
            # 格式化单元格引用
            formatted_ref = resolver.format_cell_ref(cell_ref)
            suffix = f'（{impurity_name}）' if impurity_name else ''
            errors.append(CheckError(
                cell_refs=formatted_ref,
                field='忽略不计',
                current_values=f'{num_val}%{suffix} / "忽略不计"',
                error_type='IGNORE_TERMS_WRONG',
                message=f'含量{num_val}% ≥ 报告限{threshold}%{suffix}，不应标记为"忽略不计"',
                suggestion='移除"忽略不计"标记，或确认该杂质的报告限标准'
            ))
        return errors


# ─────────────────────────────────────────────────────────────────
# 列结构识别
# ─────────────────────────────────────────────────────────────────
def find_columns(ws) -> List[Dict]:
    results = []
    for row_idx in range(1, min(35, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue
            value = str(cell_value).strip()
            if '含量' in value and '%' in value:
                name_col = None
                for offset in range(1, 6):
                    prev_cell = ws.cell(row=row_idx, column=col_idx - offset)
                    if prev_cell.value:
                        prev_val = str(prev_cell.value).strip()
                        if '杂质名称' in prev_val or prev_val == '名称':
                            name_col = col_idx - offset
                            break
                results.append({
                    'content_col': col_idx,
                    'report_col': col_idx + 1,
                    'name_col': name_col,
                    'start_row': row_idx + 1,
                })
    unique = []
    seen = set()
    for r in results:
        key = (r['content_col'], r['report_col'])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique if unique else [{'content_col': None, 'report_col': None, 'name_col': None, 'start_row': 1}]


def auto_detect_threshold_from_excel(ws) -> Optional[float]:
    keywords = [
        r'报告限[：:\\s]*([\\d.]+)%',
        r'报告阈值[：:\\s]*([\\d.]+)%',
        r'报告限值[：:\\s]*([\\d.]+)%',
        r'定量限[：:\\s]*([\\d.]+)%',
        r'report\\s*limit[：:\\s]*([\\d.]+)%',
        r'report\\s*threshold[：:\\s]*([\\d.]+)%',
    ]
    for row in ws.iter_rows(min_row=1, max_row=35):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            for kw_pattern in keywords:
                m = re.search(kw_pattern, text, re.IGNORECASE)
                if m:
                    try:
                        val = float(m.group(1))
                        if 0 < val < 100:
                            print(f"  🔍 自动识别报告限: {val}% (来源: {cell.coordinate})")
                            return val
                    except ValueError:
                        pass
    return None


# ─────────────────────────────────────────────────────────────────
# 主检查函数
# ─────────────────────────────────────────────────────────────────
def check_file(file_path: str,
               batch_format: Optional[BatchFormatValidator] = None,
               cli_threshold: Optional[float] = None,
               agent_decision: Optional[Dict] = None) -> Tuple[List[CheckError], ThresholdResolver, ProjectInfo]:
    """
    检查文件，返回 (错误列表, 阈值解析器, 项目信息)
    
    agent_decision: 可选，Agent 判断后的决策 {'project': str, 'threshold': float, 'impurity_thresholds': dict}
    """
    if batch_format is None:
        batch_format = BatchFormatValidator()

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"❌ 无法打开文件：{e}")
        return [], ThresholdResolver({}), ProjectInfo('', '', '', '', '', '', '')

    ws = wb.active

    # ── Step 1: 提取项目信息 ──
    extractor = ProjectInfoExtractor()
    project_info = extractor.extract(file_path, ws)
    print(extractor.to_prompt_context())

    sections = find_columns(ws)
    cell_resolver = MergedCellResolver(ws)

    # ── Step 2: 确定阈值 ──
    if cli_threshold is not None:
        config = {'default_threshold': cli_threshold, 'impurity_thresholds': {}}
        threshold_resolver = ThresholdResolver(config)
        print(f"  🎯 命令行指定报告限: {cli_threshold}%")
    elif agent_decision:
        # Agent 决策模式
        config = load_config(file_path)
        proj_cfg = {
            'default_threshold': agent_decision.get('threshold', 0.05),
            'impurity_thresholds': agent_decision.get('impurity_thresholds', {})
        }
        threshold_resolver = ThresholdResolver(config, proj_cfg)
        print(f"  🧠 Agent 决策: 项目={agent_decision.get('project', '未知')}, 报告限={agent_decision.get('threshold')}%")
    else:
        config = load_config(file_path)
        config = load_config(file_path)

        # 无 --agent-decision 时，打印提取的项目信息和可用配置，供 Agent 参考
        # 由 Agent 判断使用哪个项目配置，然后再次调用时通过 --agent-decision 指定
        threshold_resolver = ThresholdResolver(config)

        print()
        print("=" * 60)
        print("【项目信息提取 - 供 Agent 判断用】")
        print("=" * 60)
        print(f"项目编号: {project_info.project_code or '未提取到'}")
        print(f"检测项目: {project_info.detection_type or '未提取到'}")
        print(f"药品名: {project_info.drug_name or '未提取到'}")
        print(f"产品名: {project_info.product_name or '未提取到'}")
        print()
        print("原始文本：")
        print(project_info.raw_text)

        if config.get('project_configs'):
            print()
            print("可用项目配置：")
            for name, cfg in config['project_configs'].items():
                th = cfg.get('default_threshold', '未设')
                imps = list(cfg.get('impurity_thresholds', {}).keys())
                print(f"  · {name}: 默认报告限={th}%, 杂质专用阈值={imps}")

        print("=" * 60)
        print()
        print("【请 Agent 判断】")
        print("  请根据上方提取的项目信息（项目编号、药品名等），")
        print("  判断该文件应使用哪个项目配置。")
        print("  确认后，使用 --agent-decision 参数指定：")
        print("  Example: --agent-decision '{\"project\": \"LMS002\", \"threshold\": 0.05}'")
        print()
        print("  若无匹配的项目配置，将使用全局默认配置（default_threshold）。")
        print("=" * 60)

    batch_checker = BatchNumberChecker(batch_format)
    ignore_checker = IgnoreTermChecker(threshold_resolver)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value).strip()
            cell_ref = cell.coordinate
            batch_checker.check(cell_ref, value)
            for section in sections:
                if section['content_col'] and section['report_col']:
                    if cell.column == section['content_col'] and cell.row >= section['start_row']:
                        if cell.value is not None:
                            report_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            report_val = str(report_cell.value) if report_cell.value else ''
                            impurity_name = ''
                            if section.get('name_col'):
                                name_cell = ws.cell(row=cell.row, column=section['name_col'])
                                if name_cell.value:
                                    impurity_name = str(name_cell.value).strip()
                            ignore_checker.check(cell.value, report_val, cell_ref, impurity_name)

    wb.close()

    errors = []
    errors.extend(batch_checker.final_check(cell_resolver))
    errors.extend(ignore_checker.final_check(cell_resolver))

    return errors, threshold_resolver, project_info


# ─────────────────────────────────────────────────────────────────
# 输出格式化
# ─────────────────────────────────────────────────────────────────
def print_results(file_path: str, errors: List[CheckError], project_info: ProjectInfo):
    basename = os.path.basename(file_path)
    if not errors:
        print(f"📋 {basename}")
        print(f"  ✅ 未发现明显错误")
        return
    print(f"📋 {basename}\n  ⚠️ 发现 {len(errors)} 个潜在问题")
    print("┌──────────┬──────────────────────────────────────────────────────────────────┐")
    print("│ 位置     │ 问题                                                           │")
    print("├──────────┼──────────────────────────────────────────────────────────────────┤")
    for e in errors:
        print(f"│ {e.cell_refs:<8} │ {e.message:<64} │")
    print("└──────────┴──────────────────────────────────────────────────────────────────┘")


def print_project_info(project_info: ProjectInfo, config=None):
    """打印项目信息提取结果（供 Agent 参考）"""
    print("\n" + "=" * 60)
    print("【项目信息提取 - 供 Agent 判断用】")
    print("=" * 60)
    print(project_info.summary())
    print()
    print("提取的原始文本：")
    print(project_info.raw_text)
    
    # 输出可用配置列表，供 Agent 参考
    if config and config.get('project_configs'):
        print()
        print("可用项目配置：")
        for name, cfg in config['project_configs'].items():
            th = cfg.get('default_threshold', '未设')
            imps = list(cfg.get('impurity_thresholds', {}).keys())
            print(f"  · {name}: 默认报告限={th}%, 杂质专用阈值={imps}")
    
    print("=" * 60 + "\n")
    print("【Agent 判断提示】")
    print("  请根据上方提取的项目信息（项目编号、药品名、产品名），")
    print("  结合可用配置，判断使用哪个项目配置。")
    print("  可使用 --agent-decision 参数指定决策结果。")
    print()

# ─────────────────────────────────────────────────────────────────
# CLI 入口
# ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='HPLC有关物质检测数据明显错误检查器（AI 增强版）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
用法:
  # 普通检查（自动匹配项目）
  python3 scripts/checker.py <文件路径>

  # 指定 Agent 决策结果
  python3 scripts/checker.py <文件路径> --agent-decision '{"project": "LMS002", "threshold": 0.05}'

  # 指定全局阈值
  python3 scripts/checker.py <文件路径> --threshold 0.06

  # 仅提取项目信息（不检查）
  python3 scripts/checker.py <文件路径> --extract-only
'''
    )
    parser.add_argument('file', nargs='?', help='待检查的 Excel 文件路径')
    parser.add_argument('--threshold', '-t', type=float, default=None,
                        help='指定全局报告限阈值')
    parser.add_argument('--agent-decision', '-a', type=str, default=None,
                        help='Agent 决策结果（JSON 格式）')
    parser.add_argument('--extract-only', '-e', action='store_true',
                        help='仅提取项目信息，不执行检查')
    parser.add_argument('--version', '-v', action='version', version='%(prog)s 0.9.0')

    args = parser.parse_args()

    if not args.file:
        parser.print_help()
        sys.exit(1)

    file_path = args.file
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        sys.exit(1)

    # 解析 Agent 决策
    agent_decision = None
    if args.agent_decision:
        try:
            agent_decision = json.loads(args.agent_decision)
        except json.JSONDecodeError:
            print(f"❌ --agent-decision JSON 格式错误：{args.agent_decision}")
            sys.exit(1)

    errors, resolver, project_info = check_file(
        file_path,
        cli_threshold=args.threshold,
        agent_decision=agent_decision
    )

    if args.extract_only:
        config = checker.load_config(file_path) if os.path.exists(file_path) else {}
        print_project_info(project_info, config)
    else:
        print_results(file_path, errors, project_info)


if __name__ == '__main__':
    main()