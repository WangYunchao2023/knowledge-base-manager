#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
跟检数据错误检查引擎
版本: 0.1.0
"""

import re
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
from dataclasses import dataclass, field

# Excel 支持
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 需要 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)


@dataclass
class CheckError:
    """检查发现的错误"""
    row: int
    col: int
    cell_ref: str
    field_name: str
    current_value: str
    error_type: str
    severity: str  # ERROR / WARNING / INFO
    message: str
    suggestion: str

    def __str__(self):
        return (f"[{self.severity}] 第{self.row}行 {self.cell_ref} | {self.field_name}\n"
                f"  当前值：{self.current_value}\n"
                f"  问题：{self.message}\n"
                f"  建议：{self.suggestion}")


@dataclass
class CheckResult:
    """检查结果汇总"""
    file_path: str
    total_cells: int = 0
    errors: List[CheckError] = field(default_factory=list)
    sheet_warnings: Dict[str, int] = field(default_factory=dict)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def error_stats(self) -> Dict[str, int]:
        stats = {'ERROR': 0, 'WARNING': 0, 'INFO': 0}
        for e in self.errors:
            if e.severity in stats:
                stats[e.severity] += 1
        return stats


class BatchNumberChecker:
    """批号格式检查"""

    # 常见批号格式模式
    BATCH_PATTERNS = [
        r'^\d{8}-\d{3}$',           # 20240401-001 (日期-序号)
        r'^\d{4}-\d{2}-\d{3}$',    # 2024-04-001
        r'^[A-Z]{2}\d{6,10}$',      # AB1234567890
        r'^\d{10,12}$',             # 纯数字长批号
        r'^[A-Z]\d{4,8}[A-Z]?$',    # 字母开头+数字
    ]

    # 疑似写反的模式（如 241-01 → 2024-01）
    SUSPECTED_REVERSED = [
        (r'^(\d{2})-(\d{2})$', r'20\1-0\2'),  # 241-01 → 2024-01
        (r'^(\d{2})(\d{2})-(\d{2})$', r'20\1\2-\3'),  # 2404-01 → 2024-01
    ]

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}
        self.min_length = self.config.get('min_length', 6)
        self.max_length = self.config.get('max_length', 20)

    def check(self, value: str) -> List[CheckError]:
        """检查批号格式，返回错误列表"""
        errors = []
        if not value or not isinstance(value, str):
            return errors

        value = value.strip()
        if len(value) < self.min_length:
            errors.append(CheckError(
                row=0, col=0, cell_ref='',
                field_name='批号', current_value=value,
                error_type='BATCH_TOO_SHORT',
                severity='WARNING',
                message=f'批号过短（{len(value)}位），疑似填写不完整',
                suggestion=self._guess_batch_number(value)
            ))
        elif len(value) > self.max_length:
            errors.append(CheckError(
                row=0, col=0, cell_ref='',
                field_name='批号', current_value=value,
                error_type='BATCH_TOO_LONG',
                severity='WARNING',
                message=f'批号过长（{len(value)}位），超过常规长度',
                suggestion='请确认是否为正确的批号'
            ))

        # 检查非法字符
        illegal = re.findall(r'[\s\\|{}]', value)
        if illegal:
            errors.append(CheckError(
                row=0, col=0, cell_ref='',
                field_name='批号', current_value=value,
                error_type='BATCH_ILLEGAL_CHARS',
                severity='WARNING',
                message=f'批号包含非法字符：{illegal}',
                suggestion='批号只能包含字母、数字、连字符(-)、斜杠(/)'
            ))

        # 检查疑似写反
        for pattern, replacement in self.SUSPECTED_REVERSED:
            if re.match(pattern, value):
                fixed = re.sub(pattern, replacement, value)
                errors.append(CheckError(
                    row=0, col=0, cell_ref='',
                    field_name='批号', current_value=value,
                    error_type='BATCH_POSSIBLY_REVERSED',
                    severity='ERROR',
                    message='批号疑似写反（如 241-01 应为 2024-01）',
                    suggestion=fixed
                ))

        return errors

    def _guess_batch_number(self, value: str) -> str:
        """尝试推测正确的批号"""
        today = datetime.now()
        today_str = today.strftime('%Y%m%d')
        return f'{today_str}-XXX（请根据实际情况填写）'


class DateChecker:
    """日期格式和逻辑检查"""

    # 常见日期格式
    DATE_FORMATS = [
        '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
        '%Y%m%d', '%d-%m-%Y', '%d/%m/%Y',
    ]

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}
        self.min_year = self.config.get('min_year', 2000)
        max_offset = self.config.get('max_year_offset', 10)
        self.max_year = datetime.now().year + max_offset

    def parse_date(self, value: str) -> Tuple[Optional[datetime], str]:
        """尝试解析日期，返回 (datetime, 原始格式)"""
        if not value or not isinstance(value, str):
            return None, ''

        value = value.strip()

        # 尝试每种格式
        for fmt in self.DATE_FORMATS:
            try:
                dt = datetime.strptime(value, fmt)
                return dt, fmt
            except ValueError:
                continue

        # 尝试从混杂文本中提取
        date_patterns = [
            r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})',
            r'(\d{4})(\d{2})(\d{2})',
        ]
        for pattern in date_patterns:
            match = re.search(pattern, value)
            if match:
                try:
                    parts = match.groups()
                    dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                    return dt, 'extracted'
                except ValueError:
                    continue

        return None, ''

    def check_date_validity(self, value: str) -> Optional[CheckError]:
        """检查日期是否有效（存在）"""
        dt, fmt = self.parse_date(value)
        if dt is None:
            return None

        # 检查年份范围
        if dt.year < self.min_year or dt.year > self.max_year:
            return CheckError(
                row=0, col=0, cell_ref='',
                field_name='日期', current_value=value,
                error_type='DATE_OUT_OF_RANGE',
                severity='WARNING',
                message=f'日期 {dt.year} 年超出合理范围（{self.min_year}-{self.max_year}）',
                suggestion=f'请确认年份是否正确'
            )
        return None

    def check_format_consistency(self, values: List[str]) -> Optional[CheckError]:
        """检查一组日期值的格式是否统一"""
        if len(values) < 2:
            return None

        formats = {}
        for v in values:
            _, fmt = self.parse_date(v)
            if fmt:
                formats[fmt] = formats.get(fmt, 0) + 1

        if len(formats) > 1:
            dominant_fmt = max(formats, key=formats.get)
            return CheckError(
                row=0, col=0, cell_ref='',
                field_name='日期', current_value=', '.join(values[:3]) + ('...' if len(values)>3 else ''),
                error_type='DATE_FORMAT_INCONSISTENT',
                severity='WARNING',
                message=f'日期格式不统一，存在 {len(formats)} 种格式',
                suggestion=f'建议统一使用 {dominant_fmt} 格式'
            )
        return None

    def check_date_contradiction(self, date1: str, date2: str, label1: str = '日期1', label2: str = '日期2') -> Optional[CheckError]:
        """检查两个日期是否有逻辑矛盾"""
        dt1, _ = self.parse_date(date1)
        dt2, _ = self.parse_date(date2)
        if dt1 is None or dt2 is None:
            return None

        if dt1 > dt2:
            return CheckError(
                row=0, col=0, cell_ref='',
                field_name='日期', current_value=f'{label1}={date1}, {label2}={date2}',
                error_type='DATE_CONTRADICTION',
                severity='ERROR',
                message=f'{label1} 晚于 {label2}，存在逻辑矛盾',
                suggestion=f'请确认填写顺序'
            )
        return None


class InspectionDataChecker:
    """跟检数据综合检查器"""

    # 常见列名关键词（用于识别字段类型）
    FIELD_KEYWORDS = {
        '批号': ['批号', 'Batch', 'Lot', '批', '样号'],
        '生产日期': ['生产日期', '生产', '制造日期', ' Manufacturing Date', 'Mfg Date'],
        '取样日期': ['取样日期', '取样时间', '采样日期', 'Sampling Date', 'Sample Date'],
        '检测日期': ['检测日期', '检验日期', '分析日期', 'Test Date', 'Analysis Date'],
        '审核日期': ['审核日期', '复核日期', 'Approval Date'],
        '温度': ['温度', 'Temp', 'temperature'],
        '湿度': ['湿度', 'Humidity', 'RH'],
        '规格': ['规格', 'Spec', 'Specification'],
        '贮存条件': ['贮存条件', '贮存', '存储条件', 'Storage'],
    }

    def __init__(self, config_path: Optional[str] = None):
        self.config = self._load_config(config_path)
        self.batch_checker = BatchNumberChecker(self.config.get('batch_rules', {}))
        self.date_checker = DateChecker(self.config.get('date_rules', {}))

    def _load_config(self, config_path: Optional[str]) -> Dict:
        """加载配置文件"""
        default_config = {
            'batch_rules': {},
            'date_rules': {},
        }
        if config_path and os.path.exists(config_path):
            import yaml
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    default_config = yaml.safe_load(f)
            except ImportError:
                print("⚠️ 未安装 yaml 库，使用默认配置")
        return default_config

    def check_file(self, file_path: str) -> CheckResult:
        """检查整个Excel文件"""
        result = CheckResult(file_path=file_path)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"❌ 无法打开文件：{e}")
            return result

        all_date_values = []  # 收集所有日期值用于格式一致性检查

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            errors_in_sheet = self._check_sheet(ws, all_date_values)
            result.errors.extend(errors_in_sheet)
            result.sheet_warnings[sheet_name] = len(errors_in_sheet)
            result.total_cells += ws.max_row * ws.max_column

        wb.close()

        # 检查格式一致性（只检查日期值）
        date_fmt_error = self.date_checker.check_format_consistency(all_date_values)
        if date_fmt_error:
            result.errors.append(date_fmt_error)

        return result

    def _check_sheet(self, ws, all_date_values: List[str]) -> List[CheckError]:
        """检查单个Sheet"""
        errors = []

        # 第一遍：扫描所有单元格，识别字段类型
        cell_fields = {}  # (row, col) -> 识别的字段名
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                value = str(cell.value).strip()

                # 识别字段类型
                field_name = self._identify_field(cell.coordinate, value)
                if field_name:
                    cell_fields[cell.row, cell.column] = field_name

                # 检查批号（只有当识别为批号字段或内容符合批号特征时）
                is_batch_field = field_name == '批号' or self._is_batch_cell(cell, value)
                if is_batch_field:
                    # 排除表头行（字段名本身）
                    if not (value.lower() in ['批号', 'batch', 'lot', '样号']):
                        batch_errors = self.batch_checker.check(value)
                        for e in batch_errors:
                            e.row = cell.row
                            e.col = cell.column
                            e.cell_ref = cell.coordinate
                            errors.append(e)

                # 检查日期
                if field_name and '日期' in field_name:
                    # 排除表头
                    if not (value.lower().strip() in ['生产日期', '取样日期', '检测日期', '审核日期', '日期']):
                        all_date_values.append(value)
                        date_errors = self.date_checker.check_date_validity(value)
                        if date_errors:
                            date_errors.row = cell.row
                            date_errors.col = cell.column
                            date_errors.cell_ref = cell.coordinate
                            errors.append(date_errors)
                elif field_name is None:
                    # 没有识别为特定字段，检查是否是日期格式
                    dt, _ = self.date_checker.parse_date(value)
                    if dt:
                        all_date_values.append(value)

        # 第二遍：检查日期逻辑矛盾（如生产日期 vs 取样日期）
        self._check_date_logic(ws, errors)

        return errors

    def _identify_field(self, coord: str, value: str) -> Optional[str]:
        """根据单元格内容识别字段类型"""
        if not value:
            return None

        value_lower = value.lower()
        value_clean = value_lower.strip()

        # 排除纯表头（只有字段名，没有实际值的情况）
        # 如果单元格内容就是字段名本身（短且精确匹配），返回None
        for field_name, keywords in self.FIELD_KEYWORDS.items():
            for kw in keywords:
                kw_lower = kw.lower()
                # 精确匹配字段名（排除"批号：XXX"这种情况）
                if value_clean == kw_lower:
                    return None  # 这是表头，不算数据
                # 包含字段名关键词（可能是"批号：ABC123"）
                if kw_lower in value_clean:
                    return field_name

        return None

    def _is_batch_cell(self, cell, value: str) -> bool:
        """判断是否为批号单元格"""
        # 批号特征：字母+数字组合，长度适中
        if not value:
            return False

        # 检查是否像批号（通过上下文或格式判断）
        # 这里简化处理，实际可能需要结合表头判断
        has_letters = any(c.isalpha() for c in value)
        has_digits = any(c.isdigit() for c in value)
        length = len(value.replace('-', '').replace('/', ''))

        return has_letters and has_digits and 4 <= length <= 20

    def _check_date_logic(self, ws, errors: List[CheckError]):
        """检查日期逻辑矛盾"""
        dates_in_row = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                value = str(cell.value).strip()
                # 检查是否是日期格式（不是表头）
                dt, _ = self.date_checker.parse_date(value)
                field_name = self._identify_field(cell.coordinate, value)

                if field_name and '日期' in field_name:
                    if cell.row not in dates_in_row:
                        dates_in_row[cell.row] = {}
                    dates_in_row[cell.row][field_name] = value
                elif dt and not field_name:
                    # 可能值日期但没有被识别为日期字段的行内数据
                    if cell.row not in dates_in_row:
                        dates_in_row[cell.row] = {}
                    dates_in_row[cell.row]['检测日期'] = value

        # 检查同行内日期逻辑
        for row_num, row_dates in dates_in_row.items():
            prod_date = row_dates.get('生产日期')
            sample_date = row_dates.get('取样日期')
            test_date = row_dates.get('检测日期')

            # 生产日期 vs 取样日期
            if prod_date and sample_date:
                error = self.date_checker.check_date_contradiction(
                    sample_date, prod_date, '取样日期', '生产日期'
                )
                if error:
                    error.row = row_num
                    errors.append(error)

            # 生产日期 vs 检测日期
            if prod_date and test_date:
                error = self.date_checker.check_date_contradiction(
                    test_date, prod_date, '检测日期', '生产日期'
                )
                if error:
                    error.row = row_num
                    errors.append(error)

            # 取样日期 vs 检测日期
            if sample_date and test_date:
                error = self.date_checker.check_date_contradiction(
                    test_date, sample_date, '检测日期', '取样日期'
                )
                if error:
                    error.row = row_num
                    errors.append(error)


def print_result(result: CheckResult):
    """打印检查结果"""
    print(f"\n📋 检查报告：{os.path.basename(result.file_path)}")
    print("=" * 50)

    if result.error_count == 0:
        print(f"✅ 未发现错误（共扫描 {result.total_cells} 个单元格）")
        return

    stats = result.error_stats
    print(f"⚠️ 发现 {result.error_count} 个潜在问题")
    print(f"   ERROR: {stats['ERROR']} | WARNING: {stats['WARNING']} | INFO: {stats['INFO']}")
    print("-" * 50)

    for error in result.errors:
        print(error)
        print()


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python checker.py <Excel文件路径>")
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        sys.exit(1)

    checker = InspectionDataChecker()
    result = checker.check_file(file_path)
    print_result(result)


if __name__ == '__main__':
    main()
